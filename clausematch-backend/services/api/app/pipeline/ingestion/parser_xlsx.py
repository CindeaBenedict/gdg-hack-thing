from pathlib import Path
from openpyxl import load_workbook
import json


def parse_xlsx(path: Path) -> str:
    wb = load_workbook(str(path), data_only=True)
    data = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(["" if v is None else str(v) for v in row])
        data[sheet] = rows
    return json.dumps(data, indent=2)


